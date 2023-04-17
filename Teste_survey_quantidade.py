from time import sleep
from openpyxl import load_workbook
import PySimpleGUI as sg
import xml.etree.ElementTree as et
import random
import shutil
import os

#logando os arquivos xlsx, com um laço de tentativa
try:
    wb = load_workbook('coordenada.xlsx')
    ws = wb.active
    workbook = load_workbook('roteiro.xlsx')
    worksheet = workbook.active
except:
    sg.popup('Não tem os arquivo nescessario na pasta!\n','1-coordenada.xlsx\n','2-roteiro.xlsx',keep_on_top=True)
    
def pasta(caminho):
    pasta = caminho
    #verificar se a pasta existe se não existir ele ira criar
    if not os.path.exists(pasta):
        os.makedirs(pasta)

pasta(os.path.abspath('moradia1//'))
pasta(os.path.abspath('edificio1//'))
pasta(os.path.abspath('edificio1//edificio'))
pasta(os.path.abspath('edificio1//apartamentos//'))
pasta(os.path.abspath('moradia1//'))
                            
#laço de repetição com tempo determinado max 400 tentativa
for i in range(2,402):
    selected_theme = 'Reddit'
    sg.theme(selected_theme)
    minha_lista = []    
    sleep(.1)
    #encontrar e atribuir valores as variaveis dde uma planilha xlsx
    coordx = str(ws[f'A{i}'].value)
    coordy = str(ws[f'B{i}'].value)
    #uma variavel recebendo coordenadas, trocando ',' por '.' e concatenando com uma virgula no meio
    coordenada = coordy.replace(",",".") + ', ' + coordx.replace(",",".")
    #uma condição para que quando valor da variavel for vazio, quebre o laço
    if coordenada == 'None, None':
        break
    google_cep = str(ws[f'G{i}'].value)
    numero = str(ws[f'C{i}'].value)
    quantidade = str(ws[f'D{i}'].value)
    predio = str(ws[f'E{i}'].value)
    #uma condição para que quando valor da variavel for vazio, quebre o laço
    if google_cep == 'None':
        cep = str(ws[f'F{i}'].value)
    else:
        cep = str(ws[f'G{i}'].value)
    #encontrar e atribuir valores as variaveis dde uma planilha xlsx
    bairro = str(worksheet[f'M{i}'].value)
    roteiro = str(worksheet[f'C{i}'].value)
    localidade = str(worksheet[f'H{i}'].value)
    cod_logradouro = str(worksheet[f'N{i}'].value)
    logradouro = str(worksheet[f'Q{i}'].value) + ' ' + str(worksheet[f'O{i}'].value) + ', ' + str(worksheet[f'M{i}'].value) + ', ' + str(worksheet[f'G{i}'].value)+ ', ' + str(worksheet[f'J{i}'].value)+ ', ' + str(worksheet[f'G{i}'].value)+ ' - ' + str(worksheet[f'E{i}'].value)+ ' ' + f'({cod_logradouro})'
    
    #condição para fazer uma casa
    if quantidade == 'None':
        try:
            tree = et.parse('hp.xml')
            root = tree.getroot()    
            #encontrando e atribuindo novos valores ao xml
            for country in root.findall('enderecoEdificio'):
                country.find('logradouro').text = logradouro
                country.find('numero_fachada').text = numero
                country.find('cep').text = cep
                country.find('bairro').text = bairro
                country.find('id_roteiro').text = roteiro
                country.find('id_localidade').text = localidade
                country.find('cod_lograd').text = cod_logradouro
            #gerando um novo numero aleatório
            for linha in range (6):
                minha_lista.append(random.randint(1,9))
            num = int(''.join(map(str,minha_lista)))
            novo_numero = '20200824091321' + str(num)
            #modificar corrdenada no arquivo xml
            root.find('coordX').text = str(coordx)
            root.find('coordY').text = str(coordy)
            root.find('localidade').text = str(worksheet[f'J{i}'].value)
            #escrever xml
            tree.write('moradia1//moradia1.xml')
            #transformar arquivo zip na pasta moradia1
            shutil.make_archive(f'survey//KLAYTON_{novo_numero}','zip','./','moradia1//moradia1.xml',)
            caminho_do_arquivo = os.path.abspath('moradia1//moradia1.xml') 
            #deletar arquivo xml
            os.remove(caminho_do_arquivo)      
        except FileNotFoundError: 
            sg.popup_no_wait("Certifique-se de que o arquivo esteja na pasta:\nhp.xml\n",keep_on_top=True)
            sleep(2) 
        except PermissionError: 
            sg.popup_no_wait("Sem permissão para excluir o arquivo XML!\nmoradia1.xml",keep_on_top=True)
            sleep(2)
        except Exception as e: 
            sg.popup_no_wait("Erro ao tentar excluir o arquivo XML:\nmoradia1.xml,keep_on_top=True", e)
            sleep(2)

    #condição para fazer casa com casas secundaria ex: casa 1, casa 2, casa 3
    elif quantidade != 'None' and predio == 'None':
        try: 
            #logando os arquivos xml
            tree = et.parse('hp2.xml')
            root = tree.getroot()
            #transformando quantidade em um inteiro para iteirar no loop de repetição com tempo determinado
            quantidade = int(quantidade)
            for i in range(1,quantidade+1):
                sleep(0.1)
                #gerando um novo numero aleatório
                nova_lista=[]
                for linha in range (6):
                    nova_lista.append(random.randint(1,9))
                num = int(''.join(map(str,nova_lista)))
                novo_numero = f'20200824091322' + str(num)
                #encontrando e atribuindo novos valores ao xml
                for country in root.findall('enderecoEdificio'):
                    country.find('argumento1').text = str(i)
                    country.find('logradouro').text = logradouro
                    country.find('numero_fachada').text = numero
                    country.find('cep').text = cep
                    country.find('bairro').text = bairro
                    country.find('id_roteiro').text = roteiro
                    country.find('id_localidade').text = localidade
                    country.find('cod_lograd').text = cod_logradouro
                #encontrar e atribuir valores ao atributo do xml
                root.find('coordX').text = str(coordx) 
                root.find('coordY').text = str(coordy)
                root.find('localidade').text = str(worksheet[f'J{i}'].value) 
                sleep(0.2)    
                #escrever xml
                tree.write(f'moradia1//moradia1.xml')
                sleep(.5)
                #transformar arquivo zip na pasta moradia1
                shutil.make_archive(f'survey//KLAYTON_{novo_numero}','zip','./','moradia1//moradia1.xml',)
                #remover o arquivo moradia1.xml
                caminho_do_arquivo = os.path.abspath('moradia1//moradia1.xml') 
                #deletar arquivo xml
                os.remove(caminho_do_arquivo) 
        except FileNotFoundError:
            sg.popup_no_wait("Certifique-se de que o arquivo esteja na pasta:\nhp2.xml\n",keep_on_top=True)
            sleep(2)
        except PermissionError: 
            sg.popup_no_wait("Sem permissão para excluir o arquivo XML!\nmoradia1.xml",keep_on_top=True)
            sleep(2)
        except Exception as e: 
            sg.popup_no_wait("Erro ao tentar excluir o arquivo XML:\nmoradia1.xml,keep_on_top=True", e)
            sleep(2)
    
    #condição para fazer predio
    else:
        #logando os arquivos xml
        tree = et.parse('arquivo.xml')
        root = tree.getroot()
        tree_apartamento = et.parse('apartamento.xml') 
        root_apartamento = tree_apartamento.getroot()
        
        #gerando um novo numero aleatório
        nova_lista=[]
        for linha in range (6):
            nova_lista.append(random.randint(1,9))
        num = int(''.join(map(str,nova_lista)))
        novo_numero = f'20200824091322' + str(num)
        #encontrando e atribuindo novos valores ao xml
        for country in root.findall('enderecoEdificio'):
            country.find('logradouro').text = logradouro
            country.find('numero_fachada').text = numero
            country.find('cep').text = cep
            country.find('bairro').text = bairro
            country.find('id_roteiro').text = roteiro
            country.find('id_localidade').text = localidade
            country.find('cod_lograd').text = cod_logradouro
        #encontrar e atribuir valores ao atributo do xml
        root.find('coordX').text = str(coordx) 
        root.find('coordY').text = str(coordy)
        root.find('localidade').text = str(worksheet[f'J{i}'].value) 
        sleep(0.2)    
        #escrever xml
        tree.write(f'edificio1//edificio//edificio1.xml')
        #edição da quantidade de hp no predio
        final = int(quantidade)
        for i in range(1, final):
            tree_apartamento.write(f'edificio1//apartamentos//apartamento{i}.xml')
            xml_inserir = et.parse(f'edificio1//apartamentos//apartamento{i}.xml')
            elemento_pai_inserir = xml_inserir.getroot()
            xml_principal = et.parse(f'edificio1//edificio//edificio{i}.xml')  
            elemento_pai_princial = xml_principal.find('.//ucs')
            elemento_pai_princial.append(elemento_pai_inserir)
            xml_principal.write(f'edificio1//edificio//edificio{i+1}.xml')
            caminho_origem = f'edificio1//edificio//edificio{final}.xml'
            caminho_destino = f'edificio1//edificio.xml'
        #movendo o ultimo arquivo xml editado com a quantidade de hp para pasta edificio1
        shutil.move(caminho_origem,caminho_destino)
        #modificar edificio1.xml
        def modificar_xml():
            # Carregando o arquivo XML 
            tree = et.parse('edificio1//edificio.xml') 
            root = tree.getroot()
            #procurar elementos para alterar
            elementos_id = root.findall(".//uc/id")
            elementos_destinacao = root.findall(".//uc/destinacao")
            elementos_complemento3 = root.findall(".//uc/id_complemento3")
            elementos_argumento = root.findall(".//uc/argumento3")
            elementos_complemento4 = root.findall(".//uc/id_complemento4")
            elementos_logico = root.findall(".//uc/argumento4_logico")
            elementos_real = root.findall(".//uc/argumento4_real")
            contador = 0
            piso_real = -1
            piso_logico = 0
            #alterando esses elementos
            #id apartamento
            for i, elemento in enumerate(elementos_id):
                elemento.text = f'27385916{79+i}'
            #utilização    
            for i, elemento in enumerate(elementos_destinacao):
                elemento.text = f'RESIDENCIA'
            #não pode modificar    
            for i, elemento in enumerate(elementos_complemento3):
                elemento.text = '9'
            #numero do apartamento    
            for i, elemento in enumerate(elementos_argumento):
                if (i+1) % 4 == 1:
                    contador += 1
                elemento.text = f'{contador}0{i % 4}'
            #não pode modificar    
            for i, elemento in enumerate(elementos_complemento4):
                elemento.text = '7'
            #piso que logico
            for i, elemento in enumerate(elementos_logico):
                if (i+1) % 4 == 1:
                    piso_logico += 1
                elemento.text = f'{piso_logico}'
            #piso real    
            for i, elemento in enumerate(elementos_real):
                if (i+1) % 4 == 1:
                    piso_real += 1
                elemento.text = f'{piso_real}' 
            #escrever xml                                                                   
            tree.write('edificio1//edificio1.xml')
            
        modificar_xml()
        #gravar em zip arquivo
        shutil.make_archive(f'survey//KLAYTON_{novo_numero}','zip','./','edificio1//edificio1.xml',)
        minha_lista = []
        for linha in range (6):
            minha_lista.append(random.randint(1,9))
            num = int(''.join(map(str,minha_lista)))
            novo_numero = '20200824091321' + str(num)
        #caminho dos aquivo que serão deletados 
        for linha in range(0,final):
            caminho_do_arquivo_edificio0 = os.path.abspath(f'edificio1//edificio//edificio{linha}.xml')
            caminho_do_arquivo_apartamentos = os.path.abspath(f'edificio1//apartamentos//apartamento{linha}.xml')
            caminho_do_arquivo_edificio = os.path.abspath(f'edificio1//edificio.xml')
            caminho_do_arquivo_edificio1 = os.path.abspath(f'edificio1//edificio1.xml')
            #deletar arquivo xml
            try:
                os.remove(caminho_do_arquivo_apartamentos)
                os.remove(caminho_do_arquivo_edificio0)
                os.remove(caminho_do_arquivo_edificio)
                os.remove(caminho_do_arquivo_edificio1)      
            except:
                pass
        
sg.popup('criação concluida',keep_on_top=True) 