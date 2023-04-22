import pyautogui as pt
from pynput import mouse
import PySimpleGUI as sg
from openpyxl import load_workbook

#variavel qe define o tema
selected_theme = 'Reddit'
sg.theme(selected_theme)
wb = load_workbook('Arquivos xlsx//poste.xlsx')
ws = wb.active

def on_click(x, y,button,pressed):
    if button == mouse.Button.left and pressed:
        # Retornar False para a execução do listener de eventos
        return False
# Listener irá verificar quando o mouse clicará
with mouse.Listener(on_click=on_click) as listener:
    while True:
        # Assim que o mouse clicar, o listener irá encerrar e parar o loop
        if not listener.running:
            break
    
#definição da posição do mouse              
x , y = pt.position()
z = x - 65
x = y - 65

#inicio da criação do poste
for linha in range(2,401):
    #valores da planilha
    capacidade = str(ws[f'A{linha}'].value)
    tr = str(ws[f'B{linha}'].value)
    coordenada = str(ws[f'D{linha}'].value)
    #condição para parar de ler planilha
    if coordenada == 'None':
        break
    #condição se celula estiver vazia não atribuir valor
    elif tr == 'None':
        tr = ""
    coord = coordenada.split(",")
    #coordenada x e y
    coordx = coord[0]
    coordy = coord[1]
    
    print(f'{coordx}\t{coordy}\n{capacidade}\t{tr}')