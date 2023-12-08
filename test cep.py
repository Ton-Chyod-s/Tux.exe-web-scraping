import requests
from openpyxl import load_workbook
from tqdm import tqdm

def buscar_cep(lat, lon): 
    url = f"https://nominatim.openstreetmap.org/reverse?format=jsonv2&lat={lat}&lon={lon}" 
    response = requests.get(url) 
    if response.status_code == 200: 
        data = response.json() 
        if 'address' in data: 
            address = data['address'] 
            if 'postcode' in address: 
                return address['postcode'].replace("-", "")

wb = load_workbook('Arquivos xlsx//survey.xlsx')
ws = wb.active 

'''cep_final = buscar_cep(-27.45179655,-48.37881658)
print(cep_final) '''

for i in tqdm(range(2,402), desc="Carregando..."):
    coord = str(ws[f'D{i}'].value)
    cell1 = ws.cell(row=1, column=5)
    cell2 = ws.cell(row=1, column=1)
    cell3 = ws.cell(row=1, column=2)
    cell4 = ws.cell(row=1, column=3)
    cell5 = ws.cell(row=1, column=7)
    cell6 = ws.cell(row=1, column=8)
    cell7 = ws.cell(row=1, column=9)
    cell8 = ws.cell(row=1, column=10)

    cell = ws.cell(row=i, column=5)
    
    coordenada = coord.split(",")

    if coordenada[0] == 'None':
        break
    try:
        #coordenadas x e y
        coordy = str(coordenada[0])
        coordx = str(coordenada[1])
        cep_final = buscar_cep(coordx,coordy)
        if cep_final is not None:
            cell1.value = 'CEP GEOPY'
            cell2.value = 'NUMERO'
            cell3.value = 'QUANTIDADE'
            cell4.value = 'PREDIO'
            cell5.value = 'LOGRADOURO'
            cell6.value = 'BAIRRO'
            cell7.value = 'CIDADE'
            cell8.value = 'UF'
            cell.value = cep_final
        else:
            cell.value = 'CEP NÃO ENCONTRADO'
    except requests.exceptions.RequestException as e:
        cell.value = 'CEP NÃO ENCONTRADO'
        print(f"Erro ao buscar o CEP: {e}")
        
wb.save('Arquivos xlsx//survey.xlsx')
print('Arquivo salvo')